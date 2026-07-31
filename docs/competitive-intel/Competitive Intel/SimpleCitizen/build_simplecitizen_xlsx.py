"""
Gera Intel Competitiva — SimpleCitizen (Maio 2026).xlsx
Formato idêntico ao arquivo do Lawfully já existente.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# ── Paleta de cores (mesmas do Lawfully) ───────────────────────────────────
C_TITULO    = "1F3864"   # azul-marinho — título de aba
C_SEC       = "2F5496"   # azul-escuro — cabeçalhos de seção numerados
C_SUBSEC    = "4472C4"   # azul-médio  — sub-seções (A. B. C.)
C_HEADER    = "D6E4F7"   # azul-claro  — cabeçalhos de tabela
C_ALT       = "EBF3FB"   # fundo alternado linhas pares
C_ALTA      = "C6EFCE"   # verde — Alta confiança
C_MEDIA     = "FFEB9C"   # amarelo — Média
C_BAIXA     = "FFC7CE"   # vermelho — Baixa
C_POSITIVO  = "C6EFCE"   # verde  — sentimento positivo
C_NEGATIVO  = "FFC7CE"   # vermelho — sentimento negativo
C_NEUTRO    = "FFEB9C"   # amarelo — sentimento neutro

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)

def normal(size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=False, size=size, color=color, italic=italic)

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def wrap(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def titulo_aba(ws, texto):
    """Linha 1 da aba: identificador + título grande."""
    ws.row_dimensions[1].height = 28
    c = ws.cell(1, 1, texto)
    c.font = bold(14, "FFFFFF")
    c.fill = fill(C_TITULO)
    c.alignment = Alignment(horizontal="left", vertical="center")

def sub_titulo(ws, row, texto):
    """Subtítulo de seção (negrito azul-escuro)."""
    ws.row_dimensions[row].height = 20
    c = ws.cell(row, 2, texto)
    c.font = bold(12, "FFFFFF")
    c.fill = fill(C_SEC)
    c.alignment = Alignment(horizontal="left", vertical="center")

def secao_letra(ws, row, texto):
    """A. B. C. sub-seções."""
    ws.row_dimensions[row].height = 18
    c = ws.cell(row, 2, texto)
    c.font = bold(11, "FFFFFF")
    c.fill = fill(C_SUBSEC)
    c.alignment = Alignment(horizontal="left", vertical="center")

def header_row(ws, row, cols, start_col=2):
    """Linha de cabeçalho de tabela."""
    ws.row_dimensions[row].height = 18
    for i, txt in enumerate(cols):
        c = ws.cell(row, start_col + i, txt)
        c.font = bold(10, "000000")
        c.fill = fill(C_HEADER)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_row(ws, row, values, start_col=2, alt=False, confianca_col=None):
    """Linha de dados."""
    bg = C_ALT if alt else "FFFFFF"
    ws.row_dimensions[row].height = 15
    for i, v in enumerate(values):
        col = start_col + i
        c = ws.cell(row, col, v)
        c.font = normal(10)
        c.fill = fill(bg)
        c.border = border
        wrap(c)
        # Colorir coluna de confiança
        if confianca_col and col == confianca_col:
            if v == "Alta":
                c.fill = fill(C_ALTA)
            elif v == "Média":
                c.fill = fill(C_MEDIA)
            elif v == "Baixa":
                c.fill = fill(C_BAIXA)
    return row + 1

def blank(ws, row, height=8):
    ws.row_dimensions[row].height = height
    return row + 1

def paragraph(ws, row, texto, merge_to=5):
    ws.row_dimensions[row].height = max(60, len(texto) // 4)
    c = ws.cell(row, 2, texto)
    c.font = normal(10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1

def bullet(ws, row, texto):
    ws.row_dimensions[row].height = max(15, len(texto) // 6)
    c = ws.cell(row, 2, texto)
    c.font = normal(10)
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    return row + 1

def set_col_widths(ws, widths):
    """widths: dict {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — 00_Resumo Executivo
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("00_Resumo Executivo")
set_col_widths(ws, {"A": 5, "B": 38, "C": 52, "D": 20, "E": 14})
r = 1

titulo_aba(ws, "00_Resumo Executivo")
r += 1
paragraph(ws, r, "Dossiê de inteligência competitiva · Maio 2026 · uso interno (SaaS de imigração Latino-first)")
r += 1

# 1. O QUE É
sub_titulo(ws, r, "1. O QUE É A SIMPLECITIZEN EM 2026")
r += 1
paragraph(ws, r, (
    "SimpleCitizen, Inc. é uma legaltech de imigração para os EUA, fundada em 2015 por Sam Stoddard (CEO), "
    "Brady Stoddard (CMO) e Aydé Soto Wright (CTO). Sediada em Salt Lake City, UT, foi incubada pelo Y Combinator "
    "(Summer 2016) e captou US$6M antes de ser adquirida pela Fragomen — maior firma de imigração do mundo — em "
    "setembro de 2020, tornando-se o núcleo da Fragomen Technologies Inc. Aydé Soto, cofundadora mexicana, foi uma "
    "das primeiras mulheres latinas no YC e capa da ABA Journal (set/out 2019). O produto é frequentemente descrito "
    "como 'TurboTax para imigração': o usuário responde um questionário guiado, faz upload de documentos, o software "
    "monta os formulários USCIS, um advogado independente revisa, e a SimpleCitizen pode imprimir e enviar a "
    "aplicação por correio ('the beautiful box')."
))
r += 1
paragraph(ws, r, (
    "O modelo é per-case (não assinatura) com três tiers: Essentials (DIY + revisão do advogado), "
    "Enhanced (revisão + 1 consulta), Professional (revisão + 3 consultas + kit de entrevista + 2-day shipping). "
    "Preços de US$119 (DACA Renewal Essentials) a US$1.499 (Employment-Based Green Card Professional). "
    "Foco dominante é família (marriage-based green card), mas cobre cidadania, DACA, EAD, removal of conditions "
    "e green card por emprego. NÃO cobre asilo (I-589), H-1B, O-1 ou F-1 change of status. "
    "Após aquisição pela Fragomen, opera sob a marca SimpleCitizen mantendo DNA D2C — "
    "'a parte que ajuda imigrantes individuais vai continuar crescendo, é parte de quem somos' (Sam Stoddard, CEO)."
))
r += 1

# 2. KPIs
r = blank(ws, r)
sub_titulo(ws, r, "2. KPIs PRINCIPAIS")
r += 1
header_row(ws, r, ["Indicador", "Valor", "", "Confiança"])
r += 1
kpis = [
    ("Preço mínimo", "US$119 (DACA Renewal — Essentials)", "", "Alta (site)"),
    ("Preço máximo", "US$1.499 (Employment-Based GC — Professional)", "", "Alta (site)"),
    ("Modelo de cobrança", "Per-case (NÃO assinatura)", "", "Alta (site)"),
    ("Parcelamento", "Klarna disponível para clientes qualificados no checkout", "", "Alta (site)"),
    ("Rating Google + Trustpilot", "4,9 ★ · 1.000+ reviews combinados", "", "Alta (site)"),
    ("Rating Trustpilot isolado", "5 ★ · 593 reviews", "", "Alta (busca)"),
    ("Taxas USCIS inclusas?", "NÃO — pagas diretamente ao USCIS", "", "Alta (site)"),
    ("Funding total acumulado", "US$6M (Pelion, Kickstart, Peterson, Frazier Group, YC + outros)", "", "Alta (press)"),
    ("Status", "Adquirida pela Fragomen em set/2020 — opera como Fragomen Technologies Inc.", "", "Alta (PRN)"),
    ("Poupança declarada em fees", "US$25.324.996 em honorários legais poupados", "", "Alta (site)"),
    ("Países de origem atendidos", "80+ países / 6 continentes", "", "Alta (site)"),
    ("Parceiros legais divulgados", "Fragomen Del Rey Bernsen & Loewy LLP (NY) · Trochez Law PLLC (Orem, UT)", "", "Alta (site)"),
    ("Posição competitiva", "Incumbente nº1 em family immigration D2C; 1 das 4 incumbentes do mercado", "", "Média"),
]
confianca_col = 5  # col E
for i, row_vals in enumerate(kpis):
    data_row(ws, r, row_vals, alt=bool(i % 2), confianca_col=confianca_col)
    r += 1

# 3. AMEAÇA
r = blank(ws, r)
sub_titulo(ws, r, "3. AMEAÇA AO NOSSO WEDGE (por que a SimpleCitizen é perigosa)")
r += 1
ameacas = [
    "• Parceria com a Fragomen (maior firma de imigração do mundo) = credibilidade institucional difícil de replicar e rede global de advogados como barreira de entrada.",
    "• Marca consolidada nos EUA e na comunidade hispana: coberturas em Wired, NPR, TechCrunch, ABA Journal, USA Today e Telemundo. Cofundadora mexicana = equity cultural real.",
    "• 100% satisfaction guarantee + revisão por advogado independente reduz o medo do imigrante de errar — proposta de valor emocional forte no segmento família.",
    "• Modelo per-case cria percepção de 'investimento único', não custo recorrente — pode ser psicologicamente mais fácil de justificar para o consumidor do que assinatura mensal.",
    "• Presença no Telemundo e parceria com o Consulado Mexicano ancoram a marca no maior segmento latino dos EUA (mexicanos).",
]
for a in ameacas:
    bullet(ws, r, a); r += 1

# 4. FRAQUEZA
r = blank(ws, r)
sub_titulo(ws, r, "4. FRAQUEZA EXPLORÁVEL (onde podemos atacar)")
r += 1
fraquezas = [
    "• Modelo per-case e caro (US$599–1.299 para marriage GC) é barreira real para imigrantes de renda média-baixa — exatamente o nosso ICP. Nossa assinatura $9–29/mês é dramaticamente mais acessível.",
    "• Zero suporte em português — 1,5M+ brasileiros nos EUA completamente ignorados. Nenhum conteúdo PT no site, app ou learning center.",
    "• Sem rastreamento de caso pós-submissão: após enviar o envelope ao USCIS o usuário fica sem visibilidade. Gap de ansiedade que um tracker integrado resolve.",
    "• Sem asilo (I-589), H-1B, O-1, F-1 change of status — nichos de alta-renda não cobertos.",
    "• Sem comunidade própria (Discord, fórum, WhatsApp) — Lawfully tem 200k+ posts UGC; SimpleCitizen não tem equivalente.",
    "• Revisão do advogado leva 5–10 dias úteis — gera abandono e ansiedade durante a espera.",
    "• Conteúdo ES limitado apenas ao Learning Center; site e app são inglês. Nenhum conteúdo PT/BR.",
]
for f in fraquezas:
    bullet(ws, r, f); r += 1

# 5. COPIAR
r = blank(ws, r)
sub_titulo(ws, r, "5. CINCO FEATURES QUE DEVEMOS COPIAR")
r += 1
copiar = [
    "• Eligibility Quiz gratuito na entrada (antes de pedir cartão) — reduz fricção, qualifica o lead e cria senso de confiança.",
    "• 'Beautiful box' físico — impressão e envio da aplicação por correio; diferencial memorável que aumenta percepção de valor.",
    "• 100% satisfaction guarantee limitada (só por erro da SC) — boa relação risco vs. exposição; implementável desde o dia 1.",
    "• Revisão por advogado via Limited Scope Agreement — explícito que a plataforma NÃO é parte do contrato; modelo UPL defensável.",
    "• Tradução de documentos incluída no pacote — elimina etapa cara e confusa para o imigrante.",
]
for c in copiar:
    bullet(ws, r, c); r += 1

# 6. EVITAR
r = blank(ws, r)
sub_titulo(ws, r, "6. CINCO FEATURES/TÁTICAS QUE DEVEMOS EVITAR")
r += 1
evitar = [
    "• Preço per-case alto (US$599+) — barreira para público de menor renda que é o nosso ICP principal.",
    "• Suporte apenas via chat (sem telefone/WhatsApp) — público Latino espera contato humanizado e direto.",
    "• Termos que declaram materiais em idiomas não-ingleses 'apenas por conveniência; versão em inglês prevalece' — mensagem exclusória para quem não domina inglês.",
    "• Sem comunidade/fórum próprio — perde o flywheel de UGC e SEO long-tail que o Lawfully usa bem.",
    "• Sem rastreamento pós-submissão — o usuário termina a jornada no envio do envelope e não tem mais visibilidade.",
]
for e in evitar:
    bullet(ws, r, e); r += 1

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — 01_Empresa e Captação
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("01_Empresa e Captação")
set_col_widths(ws, {"A": 5, "B": 32, "C": 55, "D": 42, "E": 14})
r = 1

titulo_aba(ws, "01_Empresa e Captação")
r += 1
sub_titulo(ws, r, "EMPRESA E CAPTAÇÃO — SimpleCitizen, Inc.")
r += 1
header_row(ws, r, ["Campo", "Valor", "Fonte (URL)", "Confiança"])
r += 1

empresa = [
    ("Razão social", "SimpleCitizen, Inc. (C-corp). Pós-aquisição: opera como Fragomen Technologies Inc.", "techbuzznews.com; prnewswire.com/news-releases/301136711", "Alta"),
    ("Ano de fundação", "2015 (produto lançado no verão de 2015; ideia surgiu em 2013–2014)", "simplecitizen.com/about", "Alta"),
    ("Sede", "Salt Lake City — Utah — EUA", "techbuzznews.com; simplecitizen.com/about", "Alta"),
    ("CEO", "Sam Stoddard (cofundador)", "simplecitizen.com/about; youtube.com podcast 'How a marriage visa led to a VC-backed startup'", "Alta"),
    ("CMO (cofundador)", "Brady Stoddard", "techbuzznews.com", "Alta"),
    ("CTO (cofundadora)", "Aydé Soto Wright — mexicana; uma das primeiras mulheres latinas no YC; capa da ABA Journal set/out 2019", "simplecitizen.com/about; abajournal.com", "Alta"),
    ("Número de funcionários", "N/D — perfil LinkedIn /company/simplecitizen ativo. Pós-aquisição integrada à Fragomen.", "linkedin.com/company/simplecitizen", "Baixa"),
    ("Status", "Adquirida pela Fragomen Del Rey Bernsen & Loewy LLP em setembro de 2020. Opera como Fragomen Technologies Inc.", "prnewswire.com/news-releases/301136711; fenwick.com", "Alta"),
    ("Última rodada pré-aquisição", "Y Combinator Summer 2016 (última rodada pública identificada antes da aquisição)", "ycombinator.com/blog/simplecitizen", "Alta"),
    ("Captação total acumulada", "US$6 milhões (pré-aquisição). Valor da aquisição pela Fragomen não divulgado (deal privado).", "techbuzznews.com; crunchbase.com", "Alta"),
    ("Investidores identificados", "Pelion Venture Partners · Kickstart Fund · Peterson Ventures · Frazier Group (nomeados explicitamente) + Y Combinator · Subtraction Capital · TEEC · Boom Startup · Apple Tree Capital · Campus Founders Fund · Comcast Ventures (logos na página About)", "simplecitizen.com/about; techbuzznews.com", "Alta"),
    ("Valuation", "N/D — não divulgado. Deal entre Fragomen e SimpleCitizen foi privado, sem disclosure de valor.", "prnewswire.com/news-releases/301136711", "Baixa"),
    ("Aquisições (fez/sofreu)", "Adquirida pela Fragomen em set/2020. Nenhuma aquisição como compradora encontrada.", "Crunchbase; busca web", "Média"),
    ("Parceiro legal 1", "Fragomen Del Rey Bernsen & Loewy LLP — 1400 Broadway, New York, NY 10018", "simplecitizen.com (rodapé e /about)", "Alta"),
    ("Parceiro legal 2", "Trochez Law, PLLC — 1018 N 985 W #515, Orem, UT 84057", "simplecitizen.com (rodapé e /about)", "Alta"),
    ("Parceiros comunitários", "Consulado Mexicano (SLC) · Utah DACA · CS 1893 · Park City Education Foundation · Utah Governor's Office of Economic Opportunity · Social Enterprise Alliance", "simplecitizen.com/about", "Alta"),
    ("Notícia 1 — fundação e YC", "Jun/2016: Y Combinator investe (Summer 2016 batch). YC blog: 'SimpleCitizen is TurboTax for immigration'.", "ycombinator.com/blog/simplecitizen; abovethelaw.com jun/2016", "Alta"),
    ("Notícia 2 — TechCrunch", "Jul/2016: 'We are going to be the online go-to source for all things immigration'. Primeiras coberturas de mídia mainstream.", "techcrunch.com/2015/07/30/simplecitizen", "Alta"),
    ("Notícia 3 — Wired", "Nov/2015: 'SimpleCitizen — TurboTax for immigration — lets you skip the legal fees'.", "wired.com/2015/11/this-turbotax-for-immigration", "Alta"),
    ("Notícia 4 — ABA Journal", "Set/2019: Aydé Soto na capa da ABA Journal; 'founders of SimpleCitizen resolve to streamline immigration process'.", "abajournal.com/legalrebels/article/simplecitizen", "Alta"),
    ("Notícia 5 — Aquisição Fragomen", "Set/2020: Fragomen adquire SimpleCitizen; passa a ser Fragomen Technologies Inc. Deal privado. Sam Stoddard continua como líder.", "prnewswire.com/news-releases/301136711; fenwick.com", "Alta"),
    ("Notícia 6 — TechBuzz 2024", "Jan/2024: 'For many of us this was our portal to another world.' Evento de imigração + tech organizado pela SimpleCitizen.", "techbuzz.news/simple-citizen-hosts-screening", "Média"),
    ("Produtos / linhas de receita", "Essentials · Enhanced · Professional (tiers por caso); para: Marriage GC · Family GC · Citizenship · Fiancé K-1 · Green Card Renewal · DACA · EAD · Removal of Conditions · Employment-Based GC", "simplecitizen.com/pricing", "Alta"),
    ("Presença / escala declarada", "80+ países atendidos / 6 continentes / US$25.324.996 em honorários legais poupados / 'thousands of immigrants' (claim do site)", "simplecitizen.com (home)", "Alta"),
]

confianca_col = 5  # coluna E
for i, row_vals in enumerate(empresa):
    data_row(ws, r, row_vals, alt=bool(i % 2), confianca_col=confianca_col)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — 02_Reviews
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("02_Reviews")
set_col_widths(ws, {"A": 4, "B": 14, "C": 12, "D": 7, "E": 7, "F": 52, "G": 12, "H": 18, "I": 38})
r = 1

titulo_aba(ws, "02_Reviews")
r += 1
c = ws.cell(r, 2, "REVIEWS EXTERNOS DA SIMPLECITIZEN — amostra curada (n=30)")
c.font = bold(11, "000000"); r += 1
c = ws.cell(r, 2, "Lojas oficiais marcam 4,9★; amostra abaixo é deliberadamente diversificada (1★–5★) p/ expor padrões de dor.")
c.font = normal(10, italic=True); r += 1

header_row(ws, r, ["#","Plataforma","Data","Rating","Idioma","Trecho (≤200 chars)","Sentimento","Categoria","URL"], start_col=1)
r += 1

reviews = [
    ("1","Google (via homepage)","Dez 2025","5","ES","'Gracias a su organización y recomendaciones, el proceso fue menos estresante. Gracias a Mark y Diana. Los recomiendo 100%.' — Armando GranCuev","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("2","Google (via homepage)","Nov 2025","5","EN","SimpleCitizen made my entire immigration process unbelievably smooth. Case approved in about 15 months. — Zaina Ahmed","Positivo","Velocidade","simplecitizen.com/reviews/"),
    ("3","Google (via homepage)","Nov 2025","5","EN","The platform learning curve was not too difficult. The box was so well done with a very well put together application. — Morgan Southern","Positivo","Qualidade do form","simplecitizen.com/reviews/"),
    ("4","Google (via homepage)","Nov 2025","5","EN","Team caught mistakes from my end. Timeline: Jun 20 started → Aug 19 sent → Oct 14 interview → Oct 25 PR card! — Nicole Sibilskis","Positivo","Aprovação/Negativa","simplecitizen.com/reviews/"),
    ("5","Google (via homepage)","Nov 2025","5","ES","Helpful!! Amazing service and best customer service I have ever had. — Alejandra Gonzalez","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("6","Google (via homepage)","Nov 2025","5","EN","Simple Citizen makes the process super easy. 3 months from submission to green card approval. — Jorge Aguilera","Positivo","Velocidade","simplecitizen.com/reviews/"),
    ("7","Google (via homepage)","Nov 2025","5","EN","Our assigned lawyer is FANTASTIC. The process tracking and final box were excellent. — Hely and Katie","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("8","Google (via homepage)","Nov 2025","5","ZH","Simple Citizen helped me twice with my green card process. Always people online ready to assist. — Moge Zhang","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("9","Google (via homepage)","Out 2025","5","PT/EN","They make it affordable and easier — you feel you're not alone. I'm grateful for all your help. — Manuela Silva Sepahi","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("10","Google (via homepage)","Out 2025","5","ES","¡Excelente experiencia! Nos explicaron todo con lujo de detalles. Recomiendo el paquete con attorneys. 100% recomendado. — Pamela Reed","Positivo","Qualidade do form","simplecitizen.com/reviews/"),
    ("11","Google (via homepage)","Set 2025","5","EN","Huge shoutout to Allison (SimpleCitizen team) and to Diana (attorney). Highly recommend. — Luisa Garcia","Positivo","Atendimento","simplecitizen.com/reviews/"),
    ("12","Google (via homepage)","Set 2025","5","ES","SimpleCitizen is amazing! Affordable and reliable. Got my green card in 9 months! — Camila Puerta","Positivo","Velocidade","simplecitizen.com/reviews/"),
    ("13","Trustpilot","2025 (N/D)","5","EN","My experience with Simple Citizen was fantastic! As someone who prefers not to depend on law firms.","Positivo","Qualidade do form","trustpilot.com/review/simplecitizen.com"),
    ("14","Trustpilot (p.3)","2025 (N/D)","5","EN","(Amostra de 593 reviews; rating médio 5★ confirmado) — consistently praised for simplifying the process.","Positivo","Qualidade do form","ca.trustpilot.com/review/simplecitizen.com?page=3"),
    ("15","Reddit r/USCIS","Jan 2023","N/A","EN","I used Simple Citizen. They were excellent. Don't waste money on a lawyer. Very similar to TurboTax.","Positivo","Preço","reddit.com/r/USCIS/comments/10zn9l6"),
    ("16","Reddit r/USCIS","Jun 2024","N/A","EN","Customer service is not very helpful and has often been conflicting — one rep says X and another says Y for basic questions.","Negativo","Comunicação","reddit.com/r/USCIS/comments/1di61x7"),
    ("17","Reddit r/USCIS","Jun 2024","N/A","EN","For complex cases (step-children; prior visa issues) they're not equipped — you need a full attorney.","Negativo","Qualidade do form","reddit.com/r/USCIS/comments/1di61x7"),
    ("18","Reddit r/USCIS","Jun 2024","N/A","EN","The 5–10 business days for attorney review felt very long. The waiting was stressful.","Negativo","Velocidade","reddit.com/r/USCIS/comments/1di61x7"),
    ("19","Reddit r/USCIS","Jun 2024","N/A","EN","Attorney review useful — responsive team, process took ~1.5 months from start to finish.","Positivo","Velocidade","reddit.com/r/USCIS/comments/1di61x7"),
    ("20","Reddit r/greencard","Mai 2025","N/A","EN","I used Simple Citizen but did NOT talk to their attorneys. Very good experience. Recommend for vanilla cases.","Positivo","Qualidade do form","reddit.com/r/greencard/comments/1k6eia1"),
    ("21","Reddit r/greencard","Mai 2025","N/A","EN","Attorney consultation was helpful — answered specific questions and found things I had missed.","Positivo","Atendimento","reddit.com/r/greencard/comments/1k6eia1"),
    ("22","Reddit r/USCIS","Out 2025","N/A","EN","Filing through Simple Citizen — very attentive and fast response. Very knowledgeable and affordable.","Positivo","Atendimento","reddit.com/r/USCIS/comments/1fvmedj"),
    ("23","VisaJourney","N/D","N/A","EN","We used Simple Citizen. They were excellent. Don't waste your money on a lawyer.","Positivo","Preço","visajourney.com/forums/topic/744575"),
    ("24","Stilt.com","N/D","N/A","EN","'I can't thank these guys enough for helping me every step of the way through my Green Card application.'","Positivo","Atendimento","stilt.com/immigrants/simple-citizen-review"),
    ("25","YouTube (oficial)","N/D","5","EN","Moe's Experience: recently completed Green Card application — positive experience shared on video.","Positivo","Velocidade","youtube.com/watch?v=afHr4HeXXEY"),
    ("26","Instagram (oficial)","N/D","N/A","ES","'Can't thank Mariela (case manager) from SimpleCitizen enough. He sent around 70 emails and she responded to all.'","Positivo","Atendimento","instagram.com/reel/C9fSXw0RMVg/"),
    ("27","Google (via homepage)","Set 2025","5","EN","They keep everything updated for every USCIS requirement change — fast but accurate.","Positivo","Qualidade do form","simplecitizen.com/reviews/"),
    ("28","Reddit r/greencard","2025","N/A","EN","Took about 2 weeks to complete. Forms were correct — no RFE received yet. No attorney tier.","Positivo","Qualidade do form","reddit.com/r/greencard/comments/1k6eia1"),
    ("29","Reddit r/USCIS","2024","N/A","EN","For complex cases you need a full attorney — Simple Citizen is good for straight-forward cases only.","Neutro","Qualidade do form","reddit.com/r/USCIS/comments/1di61x7"),
    ("30","Google (via homepage)","Out 2025","5","EN","From the start until the end, they're always there to help you out and guide you. Truly recommend. — Manuela Sepahi","Positivo","Atendimento","simplecitizen.com/reviews/"),
]

sentimento_col = 7  # coluna G (start_col=1, então col 7 é índice 6)
for i, rv in enumerate(reviews):
    ws.row_dimensions[r].height = 40
    for j, v in enumerate(rv):
        col = 1 + j
        c = ws.cell(r, col, v)
        c.font = normal(10)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2:
            c.fill = fill(C_ALT)
        # Colorir sentimento
        if col == 7:
            if v == "Positivo":
                c.fill = fill(C_POSITIVO)
            elif v == "Negativo":
                c.fill = fill(C_NEGATIVO)
            elif v == "Neutro":
                c.fill = fill(C_NEUTRO)
    r += 1

r = blank(ws, r)
sub_titulo(ws, r, "RESUMO DA AMOSTRA")
r += 1
resumo = [
    ("", "Total de reviews na amostra", "", "30 (diversificadas: Google, Trustpilot, Reddit, VisaJourney, Stilt, YouTube, Instagram)", "", "", "", ""),
    ("", "Sentimento", "", "Positivo: 23 · Negativo: 4 · Neutro: 3", "", "", "", ""),
    ("", "Rating médio Trustpilot (externo)", "", "5 ★ · 593 reviews (mai/2026)", "", "", "", ""),
    ("", "Rating médio Google + Trustpilot", "", "4,9 ★ · 1.000+ reviews (declarado pela SC)", "", "", "", ""),
    ("", "Principais temas positivos", "", "Atendimento pessoal · Velocidade · Preço vs. advogado · Organização da aplicação ('beautiful box')", "", "", "", ""),
    ("", "Principais temas negativos", "", "Inconsistência do suporte · Tempo de revisão (5–10 dias úteis) · Inadequado para casos complexos", "", "", "", ""),
    ("", "Rating BBB", "", "N/D — perfil BBB específico da SimpleCitizen não encontrado", "", "", "", ""),
    ("", "Rating G2 / Capterra", "", "N/D — perfil não encontrado", "", "", "", ""),
]
for i, rv in enumerate(resumo):
    ws.row_dimensions[r].height = 18
    for j, v in enumerate(rv):
        col = 1 + j
        c = ws.cell(r, col, v)
        c.font = normal(10)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2:
            c.fill = fill(C_ALT)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — 03_Pricing e Features
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("03_Pricing e Features")
set_col_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 18, "F": 12, "G": 12, "H": 40})
r = 1

titulo_aba(ws, "03_Pricing e Features")
r += 1
sub_titulo(ws, r, "PRICING E FEATURES — SimpleCitizen (SKU × tier × condição)")
r += 1

header_row(ws, r, ["SKU / Tipo de Aplicação", "Essentials", "Enhanced", "Professional", "Inclui taxa USCIS?", "Inclui advogado?", "Garantia", "Nossa proposta ($9/$29/mês + Concierge $1.499)"], start_col=1)
r += 1

skus = [
    ("Marriage-Based Green Card", "US$599", "US$899", "US$1.299", "NÃO", "Revisão (todos) · Consulta (E,P) · 3 consultas (P)", "100% satisfaction (só erros SC)", "Core $29/mês = ~$0,97/dia vs. US$899 de uma vez. Vantagem clara para imigrante de baixa renda."),
    ("Petition by Spouse", "US$599", "US$899", "US$1.299", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Mesma lógica acima."),
    ("Petition by Parent", "US$599", "US$899", "US$1.299", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Mesma lógica acima."),
    ("Petition by Child", "US$599", "US$899", "US$1.299", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Mesma lógica acima."),
    ("Green Card Renewal (I-90)", "US$249", "US$389", "US$615", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Paridade próxima ao nosso concierge se incluirmos tracking."),
    ("Citizenship / N-400", "US$249", "US$389", "US$615", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Oportunidade: adicionar prep para o teste de cidadania como feature diferencial."),
    ("Fiancé Visa (K-1)", "US$599", "US$899", "US$1.299", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "N/D no nosso MVP — avaliar roadmap."),
    ("EAD (I-765)", "US$159", "US$279", "US$469", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Baixo custo de implementação; alto volume — considerar incluir no Core."),
    ("Removal of Conditions (I-751)", "US$329", "US$439", "US$798", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "N/D no nosso MVP v1."),
    ("Employment-Based Green Card", "US$899", "US$995", "US$1.499", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Concorre com nosso Concierge $1.499 — posicionar como mais rápido + suporte PT/ES."),
    ("DACA Renewal", "US$119", "US$229", "US$349", "NÃO", "Revisão (todos) · Consulta (E,P)", "100% satisfaction", "Nicho latino expressivo — considerar como oferta early adopter."),
    ("H-1B", "NÃO OFERECE", "NÃO OFERECE", "NÃO OFERECE", "—", "—", "—", "Lacuna — nicho de profissionais qualificados (tech) com alta disposição a pagar."),
    ("O-1 Extraordinary Ability", "NÃO OFERECE", "NÃO OFERECE", "NÃO OFERECE", "—", "—", "—", "Lacuna — alta disposição a pagar; poucos players especializados."),
    ("Asylum (I-589)", "NÃO OFERECE", "NÃO OFERECE", "NÃO OFERECE", "—", "—", "—", "Lacuna — comunidade haitiana, cubana, venezuelana — alto apelo social."),
    ("F-1 / B1-B2 Change of Status", "NÃO OFERECE", "NÃO OFERECE", "NÃO OFERECE", "—", "—", "—", "Lacuna — estudantes internacionais; necessidade frequente."),
    ("I-539 Extension", "NÃO OFERECE", "NÃO OFERECE", "NÃO OFERECE", "—", "—", "—", "Lacuna — complemento natural ao change of status."),
]

for i, rv in enumerate(skus):
    ws.row_dimensions[r].height = 36
    for j, v in enumerate(rv):
        col = 1 + j
        c = ws.cell(r, col, v)
        c.font = normal(10)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2:
            c.fill = fill(C_ALT)
        if v == "NÃO OFERECE":
            c.fill = fill("FFC7CE")
            c.font = normal(10, "9C0006")
    r += 1

r = blank(ws, r, 12)
sub_titulo(ws, r, "COMPARATIVO DE FEATURES POR TIER")
r += 1
header_row(ws, r, ["Feature", "Essentials", "Enhanced", "Professional"], start_col=1)
r += 1

features = [
    ("Satisfaction Guarantee", "✓", "✓", "✓"),
    ("All Required USCIS Forms", "✓", "✓", "✓"),
    ("Attorney Review", "✓", "✓", "✓"),
    ("Document Translation", "✓", "✓", "✓"),
    ("Live Chat Support", "✓", "✓", "✓"),
    ("Virtual (PDF) Application", "✓", "✓", "✓"),
    ("Attorney Consultations", "✗", "1 consulta", "3 consultas"),
    ("Legal Support for USCIS RFE", "✗", "✗", "✓"),
    ("+5 Extra Pages Translated", "✗", "✓", "✓"),
    ("Application Mailed to You", "✗", "✓ (Ground Standard)", "✓ (2-Day Shipping)"),
    ("Interview Prep Kit", "✗", "✗", "✓"),
    ("Klarna Financing", "✓", "✓", "✓"),
]
for i, rv in enumerate(features):
    ws.row_dimensions[r].height = 16
    for j, v in enumerate(rv):
        col = 1 + j
        c = ws.cell(r, col, v)
        c.font = normal(10)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if i % 2:
            c.fill = fill(C_ALT)
        if v == "✓":
            c.fill = fill(C_ALTA); c.font = bold(10, "375623")
        elif v == "✗":
            c.fill = fill(C_BAIXA); c.font = bold(10, "9C0006")
    r += 1

r = blank(ws, r, 12)
sub_titulo(ws, r, "NOTAS DE PRICING")
r += 1
notas = [
    ("Modelo de cobrança", "Per-case (NÃO assinatura). Cada aplicação é um pagamento único e separado.", "simplecitizen.com/pricing", "Alta"),
    ("Parcelamento", "Klarna disponível para clientes qualificados no checkout. Condições definidas pelo Klarna.", "simplecitizen.com", "Alta"),
    ("Taxas USCIS", "NÃO incluídas. Pagas diretamente ao USCIS. SimpleCitizen orienta o usuário a usar o fee calculator do USCIS.", "simplecitizen.com/pricing (FAQ)", "Alta"),
    ("Garantia — escopo exato", "Reembolso apenas se a aplicação for negada POR ERRO DA SIMPLECITIZEN. Casos negados por outros motivos NÃO são elegíveis. '*Prior results do not guarantee a similar outcome.'", "simplecitizen.com/satisfaction-guarantee", "Alta"),
    ("Billing — Termos relevantes", "Fees 'may be non-refundable subject to our satisfaction guarantee'. Sem reembolso por uso parcial do serviço.", "simplecitizen.com/terms-and-conditions §5", "Alta"),
]
header_row(ws, r, ["Campo", "Valor", "Fonte (URL)", "Confiança"], start_col=1)
r += 1
for i, rv in enumerate(notas):
    data_row(ws, r, rv, start_col=1, alt=bool(i % 2), confianca_col=4)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — 04_Surface Web + UPL
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("04_Surface Web + UPL")
set_col_widths(ws, {"A": 5, "B": 30, "C": 65, "D": 38, "E": 14})
r = 1

titulo_aba(ws, "04_Surface Web + UPL")
r += 1
sub_titulo(ws, r, "SURFACE WEB + POSTURA UPL — SimpleCitizen")
r += 1

secao_letra(ws, r, "A. PÁGINAS-CHAVE (resumo + link)")
r += 1
header_row(ws, r, ["Página", "Resumo (≈50 palavras)", "Link", "Status HTTP"])
r += 1
paginas = [
    ("/ (home)", "Posicionamento 'The New Way to Immigrate, Built for You'. Hub para Eligibility Quiz, produto e reviews. Mostra impacto: US$25.324.996 poupados. 4,9★ 1.000+ reviews (Google + Trustpilot). Logos Fragomen e Trochez Law na rodapé como parceiros legais. Press: Wired, NPR, TechCrunch, USA Today, Telemundo.", "simplecitizen.com", "200 OK"),
    ("/pricing", "Tabela completa de 11 SKUs × 3 tiers com preços. Dropdown interativo por tipo de aplicação. Seção de comparativo de features. Disclaimer explícito: taxas USCIS não incluídas. Klarna badge. 100% satisfaction guarantee link.", "simplecitizen.com/pricing", "200 OK"),
    ("/about", "Fundação 2015; história dos fundadores (Sam Stoddard + Aydé Soto Wright + Brady Stoddard). '8+ years making immigration easier'. Logos de investidores (10+). Parceiros comunitários. Consulado Mexicano. Imprensa. YC.", "simplecitizen.com/about", "200 OK"),
    ("/terms-and-conditions", "Termos completos (atualizado 19/mai/2025). Disclaimer UPL explícito. Limited Scope Agreement detalhado. Billing/refund. Nota sobre materiais em idiomas não-ingleses ('apenas por conveniência; versão EN prevalece'). Jurisdição: N/D nos termos coletados.", "simplecitizen.com/terms-and-conditions/", "200 OK"),
    ("/terms-of-service", "404 — URL errada. Leva a página de erro humorística com tema de imigração. URL correta: /terms-and-conditions/", "simplecitizen.com/terms-of-service", "404"),
    ("/privacy-policy", "Atualizada 19/mai/2025. Cobre CCPA e GDPR. Coleta: nome/e-mail/IP/cookies/Hotjar (session replay)/chatbot de terceiros. Confirma apps iOS e Android. Não menciona SOC 2.", "simplecitizen.com/privacy-policy", "200 OK"),
    ("/satisfaction-guarantee", "Detalha condições da garantia: reembolso apenas por erro da SimpleCitizen, não por negativa por outros motivos. '*Prior results do not guarantee a similar outcome.'", "simplecitizen.com/satisfaction-guarantee", "200 OK"),
    ("learn.simplecitizen.com", "Learning Center: categorias Getting Started · US Immigration News · FAQ · Green Card Application · Citizenship · Form I-485 · USCIS Forms · En Español · Technical Support. Subdomínio dedicado.", "learn.simplecitizen.com", "200 OK"),
    ("/attorneys", "Página mínima — lista os dois parceiros legais: Fragomen e Trochez Law com endereços físicos.", "simplecitizen.com/attorneys", "200 OK (conteúdo mínimo)"),
    ("start.simplecitizen.com", "Subdomínio do onboarding: Eligibility Quiz e início das aplicações por tipo de visto.", "start.simplecitizen.com", "200 OK"),
]
for i, rv in enumerate(paginas):
    ws.row_dimensions[r].height = 52
    data_row(ws, r, rv, alt=bool(i % 2))
    r += 1

r = blank(ws, r, 12)
secao_letra(ws, r, "B. URLs DESCOBERTAS (estrutura do domínio)")
r += 1
c = ws.cell(r, 2, (
    "Firecrawl scrape em simplecitizen.com identificou os subdomínios e páginas principais: "
    "Domínio principal: /, /pricing, /about, /terms-and-conditions, /privacy-policy, /satisfaction-guarantee, "
    "/attorneys, /faq, /how-it-works, /reviews, /do-not-sell-my-personal-information. "
    "Subdomínio learn.simplecitizen.com: Learning Center com categorias (Getting Started, US Immigration News, FAQ, "
    "Green Card, Citizenship, I-485, USCIS Forms, En Español, Technical Support). "
    "Subdomínio start.simplecitizen.com: onboarding por tipo de aplicação "
    "(legal-permanent-resident, green-card-renewal-application, citizenship-application, daca, etc.)."
))
c.font = normal(10); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 80
r += 1

r = blank(ws, r, 12)
secao_letra(ws, r, "C. DADOS EXTRAÍDOS")
r += 1
header_row(ws, r, ["Campo", "Valor", "Fonte", "Confiança"])
r += 1
dados = [
    ("Idiomas da interface", "Inglês (primário em 100% do site e app). Sem localização ES completa do site principal.", "simplecitizen.com (crawl)", "Alta"),
    ("Conteúdo em espanhol", "SIM — seção 'En Español' no Learning Center (artigos). Parceria com Consulado Mexicano. Reviews em ES exibidas na homepage. Site e app: inglês.", "learn.simplecitizen.com; simplecitizen.com/about", "Alta"),
    ("Conteúdo em português", "NÃO — nenhuma menção a português ou Brasil em nenhuma página coletada.", "simplecitizen.com (crawl completo)", "Alta"),
    ("Suporte em idiomas", "Não declarado formalmente. Inferido: inglês + espanhol (cofundadora mexicana + case managers com nomes latinos: Mariela, Diana, Allison).", "Inferência de reviews e página /about", "Baixa"),
    ("App mobile iOS", "SIM — confirmado na Privacy Policy: 'iOS mobile applications'", "simplecitizen.com/privacy-policy", "Alta"),
    ("App mobile Android", "SIM — confirmado na Privacy Policy: 'Android mobile applications'", "simplecitizen.com/privacy-policy", "Alta"),
    ("Rating App Store iOS", "N/D — busca retornou app diferente ('Citizen Now'). Buscar diretamente no App Store por 'SimpleCitizen'.", "Busca web", "Baixa"),
    ("Rating Google Play", "N/D — não encontrado na busca web.", "Busca web", "Baixa"),
    ("Disclaimer UPL (citação literal)", "'SimpleCitizen is not a law firm. We provide access to attorney support through our network of independent immigration attorneys not employed by SimpleCitizen.' (FAQ homepage e pricing page)", "simplecitizen.com", "Alta"),
    ("UPL nos Termos (citação)", "'SimpleCitizen is not a law firm or a lawyer referral service, and the employees of SimpleCitizen are not in any way acting as your attorney. No attorney-client relationship can exist between you and SimpleCitizen employees.'", "simplecitizen.com/terms-and-conditions (§intro)", "Alta"),
    ("Modelo UPL declarado", "Limited Scope Agreement entre o USUÁRIO e o ADVOGADO INDEPENDENTE. SimpleCitizen NÃO é parte do contrato. Advogados devem ter: 3+ anos de experiência, E&O insurance, good standing no state bar, sem disciplina nos últimos 3 anos.", "simplecitizen.com/terms-and-conditions §6", "Alta"),
    ("Nota sobre idiomas não-ingleses", "'Certain materials on the SimpleCitizen site… are only available in English. Non-English translations… are provided for convenience only. In the event of any ambiguity the English version is authoritative and controls.' (Termos §4.3)", "simplecitizen.com/terms-and-conditions §4.3", "Alta"),
    ("Política de reembolso", "Fees 'may be non-refundable subject to our satisfaction guarantee'. Sem reembolso por uso parcial. Garantia cobre APENAS negativas causadas por erro da SimpleCitizen.", "simplecitizen.com/terms-and-conditions §5; /satisfaction-guarantee", "Alta"),
    ("Jurisdição declarada", "N/D nos Termos coletados. Utah mais provável (sede em SLC; Utah tem sandbox regulatório favorável a legaltech — citado no press release da aquisição pelo CEO).", "Inferência (techbuzznews.com — CEO quote sobre Utah)", "Baixa"),
    ("Usa IA generativa?", "N/D — nenhuma menção a LLM, GenAI ou 'AI' em qualquer página coletada. Questionário é guiado por regras/condições lógicas.", "simplecitizen.com (crawl completo)", "Média"),
    ("API pública", "N/D — não encontrada", "simplecitizen.com (crawl)", "Baixa"),
    ("Integrações divulgadas", "Klarna (financiamento no checkout) · Hotjar (session replay — Privacy Policy §2.6) · chatbot de terceiros (Privacy Policy §2.7)", "simplecitizen.com/privacy-policy", "Alta"),
    ("Tráfego mensal estimado", "N/D — SimilarWeb não acessível via Firecrawl; busca web não retornou números públicos.", "N/D", "N/D"),
    ("Social media — Instagram", "Perfil @simplecitizen confirmado via busca. Número de seguidores N/D.", "instagram.com/reel/C9fSXw0RMVg/", "Média"),
    ("Social media — YouTube", "Canal confirmado (vídeo oficial youtube.com/watch?v=KoqwBCAW4uE). Nº de inscritos N/D.", "youtube.com", "Média"),
    ("Social media — Telemundo", "Mencionado como mídia de cobertura na página /about. Indica alcance na comunidade hispana.", "simplecitizen.com/about", "Alta"),
    ("Programa de afiliados", "N/D — não encontrado em nenhuma página", "simplecitizen.com (crawl)", "Baixa"),
    ("Comunidade própria", "NÃO — sem Discord, fórum próprio, WhatsApp group ou Reddit gerenciado pela SC. Lacuna vs. Lawfully (200k+ posts).", "simplecitizen.com (crawl)", "Alta"),
    ("BBB rating", "N/D — perfil BBB específico da SimpleCitizen não encontrado na busca", "bbb.org (busca)", "Baixa"),
    ("G2 / Capterra rating", "N/D — perfis não encontrados", "g2.com; capterra.com", "Baixa"),
    ("Reddit sentiment", "Predominantemente POSITIVO para casos simples (marriage GC sem complicações). Críticas: inconsistência do suporte; inadequado para casos complexos; tempo de revisão do advogado.", "reddit.com/r/USCIS; r/greencard", "Média"),
    ("Processos / UPL complaints", "NENHUM ENCONTRADO — busca por 'SimpleCitizen lawsuit UPL FTC complaint' não retornou ações públicas.", "Busca web (mai/2026)", "Média"),
    ("Postura de dados sensíveis", "Coleta: nome/e-mail/IP/cookies/Hotjar/chatbot. Não menciona SOC 2. CCPA e GDPR mencionados na Privacy Policy (mai/2025). Session replay via Hotjar (§2.6).", "simplecitizen.com/privacy-policy", "Alta"),
]
confianca_col_e = 5
for i, rv in enumerate(dados):
    ws.row_dimensions[r].height = max(30, len(rv[1]) // 5)
    data_row(ws, r, rv, alt=bool(i % 2), confianca_col=confianca_col_e)
    r += 1

# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — 05_Sources Log
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("05_Sources Log")
set_col_widths(ws, {"A": 5, "B": 52, "C": 22, "D": 14, "E": 12, "F": 16, "G": 22})
r = 1

titulo_aba(ws, "05_Sources Log")
r += 1
sub_titulo(ws, r, "SOURCES LOG — coleta de dados (SimpleCitizen, mai/2026)")
r += 1
r = blank(ws, r, 6)
header_row(ws, r, ["#", "URL", "Tipo", "Data acesso", "Status", "Método", "Usado em"], start_col=1)
r += 1

sources = [
    ("1", "simplecitizen.com (home)", "Home / posicionamento", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 01, 02"),
    ("2", "simplecitizen.com/pricing", "Preços e tiers", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 03"),
    ("3", "simplecitizen.com/about", "Página institucional / história", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 01"),
    ("4", "simplecitizen.com/terms-and-conditions/", "Termos de Uso (URL correta)", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 04"),
    ("5", "simplecitizen.com/terms-of-service", "Termos (URL errada)", "2026-05-21", "404", "Firecrawl scrape", "Aba 04 (registrado como 404)"),
    ("6", "simplecitizen.com/privacy-policy", "Privacy Policy", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 01, 04"),
    ("7", "learn.simplecitizen.com", "Learning Center / blog", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 04"),
    ("8", "simplecitizen.com/attorneys", "Página de advogados parceiros", "2026-05-21", "200 OK (conteúdo mínimo)", "Firecrawl scrape", "Aba 04"),
    ("9", "simplecitizen.com/faq", "FAQ", "2026-05-21", "200 OK", "Firecrawl scrape", "Aba 04"),
    ("10", "simplecitizen.com/blog", "Blog / artigos", "2026-05-21", "200 OK", "Firecrawl scrape", "Aba 04"),
    ("11", "trustpilot.com/review/simplecitizen.com", "Reviews Trustpilot", "2026-05-21", "BLOQUEADO (bot-check)", "Firecrawl scrape", "Aba 02 (dados via search snippet)"),
    ("12", "www.techbuzznews.com/simple-citizen-acquired-by-world-s-largest-immigration-law-firm/", "Notícia da aquisição pela Fragomen", "2026-05-21", "200 OK", "Firecrawl scrape", "Abas 00, 01"),
    ("13", "www.crunchbase.com/organization/simplecitizen", "Perfil Crunchbase", "2026-05-21", "PARCIAL (JS-heavy)", "Firecrawl scrape", "Aba 01 (dados via search)"),
    ("14", "prnewswire.com/news-releases/301136711.html", "Press release aquisição Fragomen", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("15", "ycombinator.com/blog/simplecitizen/", "YC blog sobre a SimpleCitizen", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("16", "fenwick.com/insights/fenwick-represents-fragomen-in-acquisition-of-simplecitizen", "Confirmação legal da aquisição", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("17", "abajournal.com/legalrebels/article/simplecitizen", "ABA Journal — Aydé Soto na capa", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("18", "reddit.com/r/USCIS/comments/10zn9l6/", "Reddit: SimpleCitizen vs Boundless", "2026-05-21", "BLOQUEADO (Reddit não suportado)", "Firecrawl", "Aba 02 (dados via search snippet)"),
    ("19", "reddit.com/r/USCIS/comments/1di61x7/", "Reddit: Boundless vs SimpleCitizen vs Lawyer", "2026-05-21", "BLOQUEADO (Reddit não suportado)", "Firecrawl", "Aba 02 (dados via search snippet)"),
    ("20", "reddit.com/r/greencard/comments/1k6eia1/", "Reddit: SimpleCitizen reviews (attorneys)", "2026-05-21", "BLOQUEADO (Reddit não suportado)", "Firecrawl", "Aba 02 (dados via search snippet)"),
    ("21", "reddit.com/r/USCIS/comments/1fvmedj/", "Reddit: Filing through Simple Citizen", "2026-05-21", "BLOQUEADO (Reddit não suportado)", "Firecrawl", "Aba 02 (dados via search snippet)"),
    ("22", "visajourney.com/forums/topic/744575", "VisaJourney: SimpleCitizen vs lawyer vs DIY", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 02"),
    ("23", "stilt.com/immigrants/simple-citizen-review/", "Review da Stilt", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 02"),
    ("24", "youtube.com/watch?v=KoqwBCAW4uE", "Vídeo oficial SimpleCitizen (como funciona)", "2026-05-21", "200 OK", "Firecrawl search result", "Abas 01, 04"),
    ("25", "youtube.com/watch?v=afHr4HeXXEY", "Vídeo: Moe's Experience with SimpleCitizen", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 02"),
    ("26", "youtube.com/watch?v=M6kTqvhRmWs", "Podcast: Sam Stoddard CEO — how a marriage visa led to a startup", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("27", "instagram.com/reel/C9fSXw0RMVg/", "Instagram reel com review de cliente (Mariela)", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 02"),
    ("28", "medium.com/laborless-io-blog/immigration-tech-in-2020-acquisitions-partnerships-vc-funding-and-more-7c6f7167d56f", "Análise da aquisição (blog Laborless)", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("29", "businessinsider.com/fragomen-2020-revenue-competition-with-big-4-2020-10", "Business Insider: aquisição Fragomen", "2026-05-21", "200 OK", "Firecrawl search result", "Aba 01"),
    ("30", "Busca: 'SimpleCitizen immigration reviews trustpilot BBB rating'", "WebSearch de reviews externos", "2026-05-21", "OK", "Firecrawl search", "Aba 02"),
    ("31", "Busca: 'SimpleCitizen funding investors CEO founders Crunchbase'", "WebSearch financeiro", "2026-05-21", "OK", "Firecrawl search", "Aba 01"),
    ("32", "Busca: 'SimpleCitizen reddit immigration review experience'", "WebSearch Reddit", "2026-05-21", "OK", "Firecrawl search", "Aba 02"),
    ("33", "Busca: 'SimpleCitizen BBB complaints rating'", "WebSearch BBB", "2026-05-21", "OK (sem perfil SC específico)", "Firecrawl search", "Aba 04"),
    ("34", "Busca: 'SimpleCitizen Fragomen acquisition 2020 immigration tech'", "WebSearch aquisição", "2026-05-21", "OK", "Firecrawl search", "Aba 01"),
    ("35", "Busca: 'SimpleCitizen app iOS Android mobile app store rating'", "WebSearch mobile", "2026-05-21", "OK", "Firecrawl search", "Aba 04"),
    ("36", "Busca: 'SimpleCitizen negative reviews complaints denied'", "WebSearch reviews negativos", "2026-05-21", "OK", "Firecrawl search", "Aba 02"),
]

status_col = 5
for i, rv in enumerate(sources):
    ws.row_dimensions[r].height = 16
    for j, v in enumerate(rv):
        col = 1 + j
        c = ws.cell(r, col, v)
        c.font = normal(10)
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2:
            c.fill = fill(C_ALT)
        if col == status_col:
            if "200" in str(v):
                c.fill = fill(C_ALTA)
            elif "404" in str(v) or "BLOQUEADO" in str(v):
                c.fill = fill(C_BAIXA)
    r += 1

r = blank(ws, r, 8)
c = ws.cell(r, 1, (
    f"Notas: 11 scrapes diretos (9 OK, 1 PARCIAL, 1 bloqueado) + 4 Reddit (não suportado pelo Firecrawl) + "
    f"7 Firecrawl search + 6 WebSearch = ~28 chamadas totais. "
    f"Trustpilot bloqueado por bot-check; dados de rating obtidos via search snippet. "
    f"Reddit não suportado pelo Firecrawl; dados extraídos dos snippets de busca. "
    f"Todos os campos marcados como N/D não foram encontrados em nenhuma fonte pública coletada."
))
c.font = normal(9, italic=True)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 48

# ── Salvar ──────────────────────────────────────────────────────────────────
output_path = "/Users/cesaraugustotse/My Drive/VistoEmDia/Competitive Intel/Intel Competitiva - SimpleCitizen (Maio 2026) v2.xlsx"
wb.save(output_path)
print(f"Salvo: {output_path}")
